from typing import List
from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.models import AttendanceLog


class ExportService:
    @staticmethod
    def export_attendance_to_excel(
        records: List[AttendanceLog], start_date: str = None, end_date: str = None
    ) -> BytesIO:
        """流式导出考勤记录到 Excel，逐行写入避免全量 DataFrame 内存占用"""
        wb = Workbook()
        ws = wb.active
        ws.title = "考勤记录"

        # 写表头
        headers = ["记录ID", "工号", "姓名", "类型", "置信度", "结果", "时间"]
        ws.append(headers)

        # 逐行写入数据
        for record in records:
            ws.append(
                [
                    record.id,
                    record.employee_id or "",
                    record.name or "",
                    "上班" if record.action_type == "CHECK_IN" else "下班",
                    f"{record.confidence:.2f}" if record.confidence else "",
                    "成功" if record.result == "SUCCESS" else "失败",
                    record.created_at.strftime("%Y-%m-%d %H:%M:%S")
                    if record.created_at
                    else "",
                ]
            )

        # 自动调整列宽
        for idx, col in enumerate(headers, 1):
            max_length = len(col)
            for row in ws.iter_rows(
                min_row=2, min_col=idx, max_col=idx, values_only=True
            ):
                cell_len = len(str(row[0])) if row[0] else 0
                if cell_len > max_length:
                    max_length = cell_len
            ws.column_dimensions[get_column_letter(idx)].width = min(max_length + 2, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def export_attendance_summary(
        records: List[AttendanceLog], start_date: str, end_date: str
    ) -> BytesIO:
        """流式导出考勤汇总到 Excel"""
        summary = {}
        for record in records:
            key = record.employee_id or "未知"
            if key not in summary:
                summary[key] = {
                    "工号": key,
                    "姓名": record.name or "未知",
                    "上班次数": 0,
                    "下班次数": 0,
                    "成功次数": 0,
                    "失败次数": 0,
                }

            if record.action_type == "CHECK_IN":
                summary[key]["上班次数"] += 1
            else:
                summary[key]["下班次数"] += 1

            if record.result == "SUCCESS":
                summary[key]["成功次数"] += 1
            else:
                summary[key]["失败次数"] += 1

        wb = Workbook()
        ws = wb.active
        ws.title = "考勤汇总"

        # 写表头
        headers = ["工号", "姓名", "上班次数", "下班次数", "成功次数", "失败次数"]
        ws.append(headers)

        # 逐行写入汇总数据
        for row_data in summary.values():
            ws.append([row_data[h] for h in headers])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output


export_service = ExportService()
